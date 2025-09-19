import openpyxl

# list of keywords in Portuguese
Privacy_List = [
    "consentimento", "políticas", "política", "regulamento", "regulamentos", "regulatório",
    "gps", "localização", "localizações", "mapa", "comportamento", "comportamental",
    "dado", "dados", "informação", "informações", "pessoal", "pessoais", "privado",
    "compartilhando", "compartilhamento", "compartilha", "segue", "seguindo", "localiza",
    "localizando", "autorização", "autorizar", "autoriza", "consentir", "consentindo",
    "permissão", "permitir", "permissões", "propaganda", "propagandas", "anúncio",
    "anúncios", "publicidade", "publicidades", "adware", "criptografar", "criptografa",
    "criptografia", "hackear", "hackeando", "hackeado", "hackeada", "hackeia",
    "inseguro", "insegura", "insegurança", "seguro", "segura", "segurança", "abusivo",
    "abusiva", "ético", "ética", "código aberto", "código livre", "protegido", "protegida",
    "proteção", "código-fonte", "confiar", "confia", "confio", "confiável", "antiético",
    "desprotegida", "desprotegido", "desproteção", "spyware", "fraude", "fraudulento",
    "engano", "enganação", "enganar", "paga", "pago", "pagar", "pagamento", "pagamentos",
    "comprar", "compra", "compro", "comprado", "comprada", "enganado", "enganada",
    "golpe", "golpista", "golpistas", "engana", "assinar", "assinatura", "assinado",
    "assinante"
]

# Categories reorganized in Portuguese
categories = {
    'politica': {
        'keywords': ["políticas", "política", "regulamento", "regulamentos", "regulatório"],
        'column': 11
    },
    'localizacao': {
        'keywords': ["gps", "localização", "localizações", "mapa", "localiza", "localizando"],
        'column': 12
    },
    'dados': {
        'keywords': ["comportamento", "comportamental", "dado", "dados",
                     "informação", "informações", "pessoal", "pessoais",
                     "privado", "compartilhando", "compartilhamento", "compartilha",
                     "segue", "seguindo"],
        'column': 13
    },
    'permissao': {
        'keywords': ["autorização", "autorizar", "autoriza", "consentimento",
                     "consentir", "consentindo", "permissão", "permitir", "permissões"],
        'column': 14
    },
    'propaganda': {
        'keywords': ["propaganda", "propagandas", "anúncio", "anúncios",
                     "publicidade", "publicidades", "adware"],
        'column': 15
    },
    'seguranca': {
        'keywords': ["criptografar", "criptografa", "criptografia", "hackear",
                     "hackeando", "hackeado", "hackeada", "hackeia", "inseguro",
                     "insegura", "insegurança", "seguro", "segura", "segurança", "spyware"],
        'column': 16
    },
    'confianca': {
        'keywords': ["abusivo", "abusiva", "ético", "ética", "código aberto",
                     "código livre", "protegido", "protegida", "proteção", "código-fonte",
                     "confiar", "confia", "confio", "confiável", "antiético",
                     "desprotegida", "desprotegido", "desproteção"],
        'column': 17
    },
    'fraude': {
        'keywords': ["fraude", "fraudulento", "engano", "enganação", "enganar",
                     "enganado", "enganada", "golpe", "golpista", "golpistas", "engana",
                     "paga", "pago", "pagar", "pagamento", "pagamentos",
                     "comprar", "compra", "compro", "comprado", "comprada",
                     "assinar", "assinatura", "assinado", "assinante"],
        'column': 18
    }
}

def classify(worksheet):
    for row in range(2, worksheet.max_row + 1):
        if worksheet.cell(row=row, column=10).value is not None:
            cell_value = str(worksheet.cell(row=row, column=10).value).lower().split()
        elif worksheet.cell(row=row, column=3).value is not None:
            cell_value = str(worksheet.cell(row=row, column=3).value).lower().split()
        else:
            continue

        # Check if the review has something to do with privacy issues
        for keyword in Privacy_List:
            if keyword in cell_value:
                # Classify the review
                for category_name, category in categories.items():
                    column = category['column']
                    for keyword2 in category['keywords']:
                        if keyword2 in cell_value:
                            worksheet.cell(row=row, column=column).value = 'Yes'
                            break
                break

# Code base
workbook = openpyxl.load_workbook('SOURCE PATH')

for name in workbook.sheetnames:
    # Select the sheet you want to read and write to
    worksheet = workbook[name]

    print('Classifying worksheet...')

    # Give the columns a title according to their classification topic    
    for category_name, category in categories.items():
        worksheet.cell(row=1, column=category['column']).value = category_name    
    
    # Classify the worksheet
    classify(worksheet)

    print('Worksheet classification done!')

# Save the changes to the workbook and close it afterwards
workbook.save('SAVE PATH')
workbook.close()
print('Classification for the whole workbook done! The workbook has been saved.')